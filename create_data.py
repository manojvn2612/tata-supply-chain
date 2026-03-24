import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import copy, re, random

# ── 1. Load existing data ─────────────────────────────────────────────────────

SOURCE = r'tata-supply-chain/Smart MRP Dummy Data_Draft.xlsx'
OUTPUT = r'tata-supply-chain/Smart MRP_StatSampled.xlsx'

# Use extended data as richer seed (16 rows)
SEED   = r'tata-supply-chain/Smart MRP Dummy Data_Draft.xlsx'

df = pd.read_excel(SEED, sheet_name='Dummy Data')

# ── 2. Parse consumption / receipt patterns into numeric arrays ───────────────

def parse_pattern(s):
    """Extract 6 monthly numbers from pattern string."""
    nums = re.findall(r'\d+', str(s))[:6]
    return [int(n) for n in nums] if len(nums) >= 6 else None

df['_cons'] = df['Consumption Pattern / Month (2025-09..2026-02)'].apply(parse_pattern)
df['_recv'] = df['Receipt Pattern / Month (2025-09..2026-02)'].apply(parse_pattern)
df['_cons_mean'] = df['_cons'].apply(lambda x: np.mean(x) if x else np.nan)
df['_cons_std']  = df['_cons'].apply(lambda x: np.std(x)  if x else np.nan)

# ── 3. Per-type statistical profiles ─────────────────────────────────────────

NUMERIC_COLS = [
    'Price/Unit', 'Stock (Unrestricted)', 'Stock (QI)', 'Stock (Blocked)',
    'MIN', 'MAX', 'Reorder Point (Current)', 'Safety Stock',
    'MOQ', 'Lot Size Value', 'Rounding Value',
    'Lead Time Supplier→Plant (Days)', 'PO Processing (Days)', 'GR Processing (Days)',
    '_cons_mean', '_cons_std'
]

profiles = {}
for mat_type, grp in df.groupby('Mat Type'):
    profiles[mat_type] = {}
    for col in NUMERIC_COLS:
        vals = grp[col].dropna().values
        if len(vals) == 0:
            continue
        profiles[mat_type][col] = {
            'mean': float(np.mean(vals)),
            'std':  max(float(np.std(vals)), float(np.mean(vals)) * 0.1),
            'min':  float(np.min(vals)),
            'max':  float(np.max(vals)),
        }

# Categorical distributions (per mat type)
CAT_COLS = ['Base UoM', 'Lot Size Proc', 'Transit Delay Scenario']
cat_profiles = {}
for mat_type, grp in df.groupby('Mat Type'):
    cat_profiles[mat_type] = {}
    for col in CAT_COLS:
        vals = grp[col].dropna().tolist()
        cat_profiles[mat_type][col] = vals  # sample uniformly

#4. Sampling helpers

def sample_int(prof, col, lo=None, hi=None):
    p = prof.get(col)
    if not p:
        return 0
    val = int(round(np.random.normal(p['mean'], p['std'])))
    val = max(val, int(p['min']))
    val = min(val, int(p['max']))
    if lo is not None: val = max(val, lo)
    if hi is not None: val = min(val, hi)
    return val

def sample_cat(cat_prof, mat_type, col):
    opts = cat_prof.get(mat_type, {}).get(col, [])
    return random.choice(opts) if opts else ''

# Material counters (continue from existing)
mat_counters = {'FERT': 4, 'HALB': 4, 'ROH': 11}

suppliers = ['Vendor A','Vendor B','Vendor C','Vendor D','Vendor E',
             'Vendor F','Vendor G','Vendor H','Vendor I','Vendor K',
             'Vendor L','Vendor M','Vendor-managed']

descriptions = {
    'FERT': ['Brushless DC Motor Unit','Variable Frequency Drive','Linear Actuator Module',
             'Stepper Motor Assembly','Hydraulic Control Unit','AC Servo Controller',
             'Spindle Drive Module','Traction Motor Pack'],
    'HALB': ['Gate Driver PCB','Capacitor Bank Assembly','Busbar Sub-Assembly',
             'Filter Inductor Module','Cooling Plate Sub-Assembly','Control Board Assembly',
             'Power Stage Sub-Assembly','Sensor Fusion Board'],
    'ROH':  ['Silicon Carbide Wafer','Ferrite Core Ring','Electrolytic Capacitor',
             'Film Capacitor Pack','Gate Resistor Kit','Current Sensor Module',
             'Voltage Divider Assembly','Toroidal Transformer Core',
             'Shunt Resistor Array','EMI Filter Component','SMD Component Kit']
}

def new_material_id(mat_type):
    if mat_type == 'FERT':
        code = f"FG-1000{mat_counters['FERT']:02d}"
        mat_counters['FERT'] += 1
    elif mat_type == 'HALB':
        code = f"SFG-1500{mat_counters['HALB']:02d}"
        mat_counters['HALB'] += 1
    else:
        code = f"RM-2000{mat_counters['ROH']:02d}"
        mat_counters['ROH'] += 1
    return code

def generate_consumption(mean, std, uom, bom_qty=1):
    """Generate 6 months of consumption with mild trend + noise."""
    base = max(10, int(mean))
    months = []
    for i in range(6):
        trend = 1 + 0.03 * i  # ~3% monthly growth trend
        noise = np.random.normal(0, max(1, std * 0.3))
        months.append(max(1, int(base * trend + noise)))
    label = ", ".join(str(m) for m in months)
    if bom_qty > 1:
        label += f" (= parent consumption ×{bom_qty})"
    return label, months

def generate_receipt(cons_months, jitter=0.05):
    """Receipts lag consumption slightly with small noise."""
    months = [max(1, int(c * np.random.uniform(0.92, 1.08))) for c in cons_months]
    return ", ".join(str(m) for m in months)

def rounding_profile(mat_type, uom):
    if uom == 'KG': return 'RP-RM-KG'
    if mat_type == 'FERT': return 'RP-FG-EA'
    if mat_type == 'HALB': return 'RP-SFG-10'
    return 'RP-RM-EA'

def std_price_ref(mat_type):
    return 'S (Std)' if mat_type in ('FERT','HALB') else 'MAP (Moving Avg)'

def po_text(mat_id, stock, safety, cons_mean, lead, moq):
    """Generate plausible open PO text."""
    qty = max(moq, int((safety * 2 - stock) / moq) * moq)
    qty = max(qty, moq)
    po_num = random.randint(4500015000, 4500019999)
    days_out = random.randint(10, 45)
    month = '03' if days_out < 20 else '04'
    day = random.randint(1, 28)
    result = f"PO#{po_num}: {qty} EA due 2026-{month}-{day:02d}"
    if random.random() > 0.5:
        po2 = random.randint(4500015000, 4500019999)
        qty2 = max(moq, qty // 2)
        result += f";\nPO#{po2}: {qty2} EA due 2026-{month}-{min(28,day+14):02d}"
    return result

# ── 5. Generate N new rows ────────────────────────────────────────────────────

def generate_row(mat_type, bom_parent='(self)', bom_qty=1):
    prof = profiles.get(mat_type, {})
    uom  = sample_cat(cat_profiles, mat_type, 'Base UoM')
    
    price      = sample_int(prof, 'Price/Unit', lo=100)
    cons_mean  = sample_int(prof, '_cons_mean', lo=10)
    cons_std   = sample_int(prof, '_cons_std',  lo=1)
    
    cons_label, cons_months = generate_consumption(cons_mean * bom_qty, cons_std, uom, bom_qty)
    recv_label = generate_receipt(cons_months)
    
    stock_u    = sample_int(prof, 'Stock (Unrestricted)', lo=0)
    stock_qi   = sample_int(prof, 'Stock (QI)',           lo=0)
    stock_b    = sample_int(prof, 'Stock (Blocked)',      lo=0)
    
    s_stock    = sample_int(prof, 'Safety Stock',  lo=5)
    rop        = sample_int(prof, 'Reorder Point (Current)', lo=s_stock)
    min_val    = max(1, rop // 2)
    max_val    = sample_int(prof, 'MAX', lo=rop * 2)
    moq        = sample_int(prof, 'MOQ', lo=1)
    lot_proc   = sample_cat(cat_profiles, mat_type, 'Lot Size Proc')
    lot_val    = sample_int(prof, 'Lot Size Value', lo=moq)
    rnd_val    = sample_int(prof, 'Rounding Value', lo=1)
    
    lead       = sample_int(prof, 'Lead Time Supplier→Plant (Days)', lo=5)
    po_proc    = sample_int(prof, 'PO Processing (Days)', lo=1)
    gr_proc    = sample_int(prof, 'GR Processing (Days)', lo=1)
    delay      = sample_cat(cat_profiles, mat_type, 'Transit Delay Scenario')
    
    mat_id     = new_material_id(mat_type)
    desc_opts  = descriptions.get(mat_type, ['Generic Material'])
    used       = df['Description'].tolist()
    available  = [d for d in desc_opts if d not in used]
    desc       = random.choice(available) if available else random.choice(desc_opts) + f" {mat_counters[mat_type]}"
    
    supplier   = 'Vendor-managed' if mat_type == 'FERT' else random.choice(suppliers[:-1])
    po_text_   = '(n/a)' if mat_type == 'FERT' else po_text(mat_id, stock_u, s_stock, cons_mean, lead, moq)
    bom_comp   = '(n/a)' if mat_type == 'ROH' else f"RM-auto:{random.randint(1,3)} EA;\nRM-auto2:{round(random.uniform(0.1,1.0),1)} KG"
    
    if mat_type == 'ROH' and bom_qty > 1:
        cons_label = cons_label  # already annotated

    return [
        "1000", mat_id, mat_type, desc, uom if uom else 'EA',
        price, std_price_ref(mat_type),
        bom_parent, bom_comp,
        cons_label, recv_label,
        stock_u, stock_qi, stock_b, po_text_,
        min_val, max_val, rop, s_stock,
        moq, lot_proc, lot_val,
        rounding_profile(mat_type, uom if uom else 'EA'), rnd_val,
        supplier, lead, po_proc, gr_proc, delay
    ]

# Generate: 3 FG, 3 SFG (children of FG), 6 RM
N_FG, N_SFG, N_RM = 10, 10, 10
generated = []

for _ in range(N_FG):
    generated.append(('FERT', generate_row('FERT')))

for i in range(N_SFG):
    parent_id = generated[i][1][1]  # Material ID of FG
    generated.append(('HALB', generate_row('HALB', bom_parent=parent_id)))

for _ in range(N_RM):
    bom_qty = random.choice([1, 2, 3])
    generated.append(('ROH', generate_row('ROH', bom_parent='FG-auto', bom_qty=bom_qty)))

# ── 6. Append to workbook ─────────────────────────────────────────────────────

wb = load_workbook(SEED)
ws = wb['Dummy Data']

def copy_row_format(ws, src_row, dst_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font      = copy.copy(src.font)
            dst.fill      = copy.copy(src.fill)
            dst.border    = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)

last_row = ws.max_row
for i, (mat_type, row_data) in enumerate(generated):
    dst_row = last_row + 1 + i
    # Use matching existing row as format template (FERT=row2, HALB=row3, ROH=row4)
    src_map = {'FERT': 2, 'HALB': 3, 'ROH': 4}
    copy_row_format(ws, src_map[mat_type], dst_row)
    for col_idx, val in enumerate(row_data, start=1):
        ws.cell(row=dst_row, column=col_idx).value = val

wb.save(OUTPUT)
print(f"✅ Generated {len(generated)} new rows  |  Total rows now: {last_row - 1 + len(generated)}")
print("\nSample generated row (FERT):")
print(generated[0][1])